"""
export_engine.py  — PDF ve Excel dışa aktarım motoru
Türkçe karakter desteği: DejaVu Sans TTF
Hücre taşması: Paragraph + wrapping
"""

import os
import sys
from datetime import datetime
import database as db

import schedule_config as _sc

def _get_days():  return _sc.get_active_days()
def _get_hours(): return _sc.get_hours_list()

DAYS      = [0, 1, 2, 3, 4]
HOURS     = list(range(1, 9))
DAY_NAMES = _sc.ALL_DAY_NAMES

# Font dosya yolu — proje klasöründe assets/fonts/ altında
_BASE   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_FONTS  = os.path.join(_BASE, "assets", "fonts")
FONT_REG  = os.path.join(_FONTS, "DejaVuSans.ttf")
FONT_BOLD = os.path.join(_FONTS, "DejaVuSans-Bold.ttf")

# Windows'ta da çalışması için system font fallback
if not os.path.exists(FONT_REG):
    for candidate in [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    ]:
        if os.path.exists(candidate):
            FONT_REG = candidate
            break

if not os.path.exists(FONT_BOLD):
    for candidate in [
        r"C:\Windows\Fonts\arialbd.ttf",
        r"C:\Windows\Fonts\ArialBD.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ]:
        if os.path.exists(candidate):
            FONT_BOLD = candidate
            break


def _register_fonts():
    """ReportLab'a Unicode fontları kaydet (ilk çağrıda bir kez yapılır)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    if "DejaVu" not in pdfmetrics._fonts:
        pdfmetrics.registerFont(TTFont("DejaVu",     FONT_REG))
        pdfmetrics.registerFont(TTFont("DejaVu-Bold", FONT_BOLD))
        from reportlab.pdfbase.pdfmetrics import registerFontFamily
        registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold",
                           italic="DejaVu", boldItalic="DejaVu-Bold")


# ─────────────────────────────────────────────────────────────────────────────
#  Ortak veri toplama
# ─────────────────────────────────────────────────────────────────────────────

def build_class_grid(class_id):
    rows = db.get_timetable(class_id)
    grid = {}
    for r in rows:
        r = dict(r)
        grid[(r['day'], r['hour'])] = {
            'subject':   r['subject_name'],
            'teacher':   r.get('teacher_name', '—'),
            'classroom': r.get('classroom_name', '—'),
        }
    return grid


def build_teacher_grid(teacher_id):
    rows = db.get_timetable_for_teacher(teacher_id)
    grid = {}
    for r in rows:
        r = dict(r)
        grid[(r['day'], r['hour'])] = {
            'subject':   r['subject_name'],
            'class':     r.get('class_name', '—'),
            'classroom': r.get('classroom_name', '—'),
        }
    return grid


# ─────────────────────────────────────────────────────────────────────────────
#  PDF — ReportLab  (Paragraph ile hücre taşması yok, DejaVu ile Türkçe OK)
# ─────────────────────────────────────────────────────────────────────────────

def _make_styles():
    _register_fonts()
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle

    hdr_style = ParagraphStyle(
        "hdr", fontName="DejaVu-Bold", fontSize=9,
        textColor=colors.white,
        alignment=1, leading=12, wordWrap="CJK")

    subj_style = ParagraphStyle(
        "subj", fontName="DejaVu-Bold", fontSize=8,
        textColor=colors.HexColor("#1a1a1a"),
        alignment=1, leading=11, wordWrap="CJK")

    sub_style = ParagraphStyle(
        "sub", fontName="DejaVu", fontSize=7,
        textColor=colors.HexColor("#3a7bd5"),
        alignment=1, leading=10, wordWrap="CJK")

    hour_style = ParagraphStyle(
        "hour", fontName="DejaVu-Bold", fontSize=8,
        textColor=colors.HexColor("#aaaaaa"),
        alignment=1, leading=11)

    empty_style = ParagraphStyle(
        "empty", fontName="DejaVu", fontSize=9,
        textColor=colors.HexColor("#cccccc"),
        alignment=1, leading=11)

    return hdr_style, subj_style, sub_style, hour_style, empty_style


def _pdf_timetable_table(grid, mode='class'):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.units import cm

    _register_fonts()
    hdr_style, subj_style, sub_style, hour_style, empty_style = _make_styles()

    HEADER_BG  = colors.HexColor("#1e3a5f")
    HOUR_BG    = colors.HexColor("#2d2d2d")
    CELL_BG    = colors.HexColor("#ebf3fb")
    EMPTY_BG   = colors.HexColor("#f8f8f8")
    BORDER_CLR = colors.HexColor("#c8c8c8")

    # Dinamik gün ve saat listesi
    DAYS  = _get_days()
    HOURS = _get_hours()

    # Başlık satırı
    header_row = [Paragraph("Saat", hdr_style)] + \
                 [Paragraph(DAY_NAMES.get(d, str(d)), hdr_style) for d in DAYS]
    data = [header_row]

    for hour in HOURS:
        row = [Paragraph(_sc.get_period_label(hour), hour_style)]
        for day in DAYS:
            entry = grid.get((day, hour))
            if entry:
                subj = entry['subject']
                sub  = entry.get('teacher' if mode == 'class' else 'class', '—')
                cell = [Paragraph(subj, subj_style),
                        Paragraph(sub,  sub_style)]
                row.append(cell)
            else:
                row.append(Paragraph("—", empty_style))
        data.append(row)

    col_widths  = [1.8*cm] + [3.8*cm]*5
    row_heights = [1.0*cm] + [1.8*cm]*len(HOURS)

    t = Table(data, colWidths=col_widths, rowHeights=row_heights,
              repeatRows=1)

    style = TableStyle([
        # Genel
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('GRID',        (0, 0), (-1, -1), 0.5, BORDER_CLR),
        ('LEFTPADDING',  (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        # Başlık
        ('BACKGROUND',  (0, 0), (-1, 0),  HEADER_BG),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [CELL_BG, colors.white]),
        # Saat sütunu
        ('BACKGROUND',  (0, 1), (0, -1), HOUR_BG),
    ])
    t.setStyle(style)
    return t


def _get_school_header():
    """Veritabanından okul bilgilerini alır, tek satır string döndürür."""
    import database as db
    s = db.get_settings()
    parts = []
    if s.get("school_name"):
        parts.append(s["school_name"])
    if s.get("academic_year"):
        parts.append(f"Öğretim Yılı: {s['academic_year']}")
    if s.get("principal"):
        parts.append(f"Müdür: {s['principal']}")
    if s.get("vice_principal"):
        parts.append(f"Müd. Yrd.: {s['vice_principal']}")
    return "  |  ".join(parts) if parts else ""


def _build_title_paragraphs(title_text, subtitle_text):
    _register_fonts()
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, HRFlowable
    from reportlab.lib.units import cm

    school_header = _get_school_header()

    school_s = ParagraphStyle(
        "pschool", fontName="DejaVu-Bold", fontSize=10,
        textColor=colors.HexColor("#1e3a5f"), spaceAfter=2, leading=14)
    title_s = ParagraphStyle(
        "ptitle", fontName="DejaVu-Bold", fontSize=15,
        textColor=colors.HexColor("#1e3a5f"), spaceAfter=2, leading=18)
    sub_s = ParagraphStyle(
        "psub", fontName="DejaVu", fontSize=9,
        textColor=colors.HexColor("#666666"), spaceAfter=8, leading=12)

    result = []
    if school_header:
        result.append(Paragraph(school_header, school_s))
        result.append(HRFlowable(width="100%", thickness=0.5,
                                  color=colors.HexColor("#c0d0e0"),
                                  spaceAfter=4))
    result.append(Paragraph(title_text, title_s))
    result.append(Paragraph(subtitle_text, sub_s))
    return result


def _make_doc(output_path):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate
    from reportlab.lib.units import cm
    return SimpleDocTemplate(
        output_path, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)


# ── Tek sınıf ──────────────────────────────────────────────────────────────
def export_pdf_class(class_id, output_path):
    with db.get_connection() as conn:
        cls = dict(conn.execute(
            "SELECT * FROM classes WHERE id=?", (class_id,)).fetchone())
    cname = f"{cls['level']}-{cls['section']}"

    doc   = _make_doc(output_path)
    story = _build_title_paragraphs(
        f"{cname} Sınıfı — Haftalık Ders Programı",
        f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"Öğrenci sayısı: {cls['student_count']}"
    )
    story.append(_pdf_timetable_table(build_class_grid(class_id), 'class'))
    doc.build(story)


# ── Tek öğretmen ───────────────────────────────────────────────────────────
def export_pdf_teacher(teacher_id, output_path):
    with db.get_connection() as conn:
        t = dict(conn.execute(
            "SELECT * FROM teachers WHERE id=?", (teacher_id,)).fetchone())
    name = f"{t['name']} {t['surname']}"

    doc   = _make_doc(output_path)
    story = _build_title_paragraphs(
        f"{name} — Haftalık Ders Programı",
        f"Branş: {t['branch']}  |  "
        f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    story.append(_pdf_timetable_table(build_teacher_grid(teacher_id), 'teacher'))
    doc.build(story)


# ── Tüm sınıflar ───────────────────────────────────────────────────────────
def export_pdf_all_classes(output_path):
    from reportlab.platypus import PageBreak
    classes = db.get_classes()
    doc   = _make_doc(output_path)
    story = []
    for i, cls in enumerate(classes):
        cls = dict(cls)
        cname = f"{cls['level']}-{cls['section']}"
        story += _build_title_paragraphs(
            f"{cname} Sınıfı — Haftalık Ders Programı",
            f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
            f"Öğrenci: {cls['student_count']}")
        story.append(_pdf_timetable_table(build_class_grid(cls['id']), 'class'))
        if i < len(classes) - 1:
            story.append(PageBreak())
    doc.build(story)


# ── Tüm öğretmenler ────────────────────────────────────────────────────────
def export_pdf_all_teachers(output_path):
    from reportlab.platypus import PageBreak
    teachers = db.get_teachers()
    doc   = _make_doc(output_path)
    story = []
    for i, t in enumerate(teachers):
        t = dict(t)
        name = f"{t['name']} {t['surname']}"
        story += _build_title_paragraphs(
            f"{name} — Haftalık Ders Programı",
            f"Branş: {t['branch']}  |  "
            f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        story.append(_pdf_timetable_table(build_teacher_grid(t['id']), 'teacher'))
        if i < len(teachers) - 1:
            story.append(PageBreak())
    doc.build(story)


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL — openpyxl
# ─────────────────────────────────────────────────────────────────────────────

def _apply_excel_grid(ws, grid, mode, title, subtitle):
    import database as db
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_BG = "1E3A5F"
    HOUR_BG   = "2D2D2D"
    CELL_ODD  = "EBF3FB"
    CELL_EVEN = "FFFFFF"
    EMPTY_BG  = "F8F8F8"
    WHITE     = "FFFFFF"
    DARK      = "1A1A1A"
    GRAY      = "AAAAAA"

    thin   = Side(style='thin',   color="C8C8C8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Satır 1: Okul adı ──────────────────────────────────────────────────
    school_info = _get_school_header()
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value     = school_info if school_info else title
    c.font      = Font(name='Arial', bold=True, size=13, color=HEADER_BG)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill      = PatternFill("solid", fgColor="E8F0FA")
    ws.row_dimensions[1].height = 26

    # ── Satır 2: Rapor başlığı ──────────────────────────────────────────────
    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value     = title
    c.font      = Font(name='Arial', bold=True, size=12, color=HEADER_BG)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── Satır 3: Alt bilgi ──────────────────────────────────────────────────
    ws.merge_cells('A3:G3')
    c = ws['A3']
    c.value     = subtitle
    c.font      = Font(name='Arial', size=9, color="888888")
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 6

    # ── Satır 5: Gün başlıkları ─────────────────────────────────────────────
    DAYS  = _get_days()
    HOURS = _get_hours()
    ws.column_dimensions['A'].width = 14
    headers = ["Saat"] + [DAY_NAMES.get(d, str(d)) for d in DAYS]
    for ci, h in enumerate(headers):
        col = get_column_letter(ci + 1)
        ws.column_dimensions[col].width = 22 if ci > 0 else 12
        c = ws.cell(row=5, column=ci+1, value=h)
        c.font      = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        c.border    = border
    ws.row_dimensions[5].height = 22

    # ── Ders satırları ──────────────────────────────────────────────────────
    for hi, hour in enumerate(HOURS):
        row_num = hi + 6
        ws.row_dimensions[row_num].height = 44

        c = ws.cell(row=row_num, column=1, value=_sc.get_period_label(hour))
        c.font      = Font(name='Arial', bold=True, size=9, color=GRAY)
        c.fill      = PatternFill("solid", fgColor=HOUR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = border

        for di, day in enumerate(DAYS):
            col_num = di + 2
            entry   = grid.get((day, hour))
            c = ws.cell(row=row_num, column=col_num)
            c.border    = border
            c.alignment = Alignment(horizontal='center', vertical='center',
                                     wrap_text=True)
            if entry:
                subj = entry['subject']
                sub  = entry.get('teacher' if mode == 'class' else 'class', '—')
                c.value = f"{subj}\n{sub}"
                c.font  = Font(name='Arial', size=9, color=DARK)
                c.fill  = PatternFill("solid", fgColor=CELL_ODD if hi%2==0 else CELL_EVEN)
            else:
                c.value = ""
                c.fill  = PatternFill("solid", fgColor=EMPTY_BG)


def export_excel_class(class_id, output_path):
    from openpyxl import Workbook
    with db.get_connection() as conn:
        cls = dict(conn.execute(
            "SELECT * FROM classes WHERE id=?", (class_id,)).fetchone())
    cname = f"{cls['level']}-{cls['section']}"
    wb = Workbook(); ws = wb.active
    ws.title = cname
    ws.sheet_view.showGridLines = False
    _apply_excel_grid(ws, build_class_grid(class_id), 'class',
        title=f"{cname} Sınıfı — Haftalık Ders Programı",
        subtitle=f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
                 f"Öğrenci: {cls['student_count']}")
    wb.save(output_path)


def export_excel_teacher(teacher_id, output_path):
    from openpyxl import Workbook
    with db.get_connection() as conn:
        t = dict(conn.execute(
            "SELECT * FROM teachers WHERE id=?", (teacher_id,)).fetchone())
    name = f"{t['name']} {t['surname']}"
    wb = Workbook(); ws = wb.active
    ws.title = name[:31]
    ws.sheet_view.showGridLines = False
    _apply_excel_grid(ws, build_teacher_grid(teacher_id), 'teacher',
        title=f"{name} — Haftalık Ders Programı",
        subtitle=f"Branş: {t['branch']}  |  "
                 f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    wb.save(output_path)


def export_excel_all_classes(output_path):
    from openpyxl import Workbook
    classes = db.get_classes()
    wb = Workbook(); wb.remove(wb.active)
    for cls in classes:
        cls = dict(cls)
        cname = f"{cls['level']}-{cls['section']}"
        ws = wb.create_sheet(title=cname)
        ws.sheet_view.showGridLines = False
        _apply_excel_grid(ws, build_class_grid(cls['id']), 'class',
            title=f"{cname} Sınıfı — Haftalık Ders Programı",
            subtitle=f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
                     f"Öğrenci: {cls['student_count']}")
    wb.save(output_path)


def export_excel_all_teachers(output_path):
    from openpyxl import Workbook
    teachers = db.get_teachers()
    wb = Workbook(); wb.remove(wb.active)
    for t in teachers:
        t = dict(t)
        name = f"{t['name']} {t['surname']}"
        ws = wb.create_sheet(title=name[:31])
        ws.sheet_view.showGridLines = False
        _apply_excel_grid(ws, build_teacher_grid(t['id']), 'teacher',
            title=f"{name} — Haftalık Ders Programı",
            subtitle=f"Branş: {t['branch']}  |  "
                     f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    wb.save(output_path)
